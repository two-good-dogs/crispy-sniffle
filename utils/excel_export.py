"""
AuditIQ — Excel Export Utility
Generates a styled, multi-sheet workbook from current dashboard data.
"""
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter


# ── Palette ───────────────────────────────────────────────────────────────────

DARK_BG    = "1F2937"
MID_BG     = "374151"
ACCENT     = "1E40AF"
GREEN      = "166534"
TEXT_WHITE = "FFFFFF"
TEXT_GRAY  = "6B7280"
BORDER_CLR = "E5E7EB"

# Sheet names — single source of truth (imported by header.py)
ALL_SHEETS = [
    "Summary Dashboard",
    "Portfolio Overview",
    "Issue Tracker",
    "Assurance Summary",
    "Adjustment Workflow",
]

# ── Module-level style singletons ─────────────────────────────────────────────

THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_CLR),
    right=Side(style="thin", color=BORDER_CLR),
    top=Side(style="thin", color=BORDER_CLR),
    bottom=Side(style="thin", color=BORDER_CLR),
)
FILL_DARK       = PatternFill("solid", fgColor=DARK_BG)
FILL_MID        = PatternFill("solid", fgColor=MID_BG)
FILL_KPI        = PatternFill("solid", fgColor="F8FAFC")
ALIGN_CENTER    = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT      = Alignment(vertical="center", indent=1)
ALIGN_HDR_WRAP  = Alignment(horizontal="center", vertical="center", wrap_text=True)
FONT_BODY       = Font(size=9)
FONT_HEADER     = Font(bold=True, color=TEXT_WHITE, size=10)

RATING_FILLS = {
    "SAT":   PatternFill("solid", fgColor="D1FAE5"),
    "RI":    PatternFill("solid", fgColor="FEF3C7"),
    "UNSAT": PatternFill("solid", fgColor="FEE2E2"),
    "NA":    PatternFill("solid", fgColor="F3F4F6"),
}
RATING_FONTS = {
    "SAT":   Font(bold=True, color="065F46"),
    "RI":    Font(bold=True, color="92400E"),
    "UNSAT": Font(bold=True, color="991B1B"),
    "NA":    Font(bold=True, color=TEXT_GRAY),
}
STATUS_FILLS = {
    "Complete":    PatternFill("solid", fgColor="D1FAE5"),
    "In Progress": PatternFill("solid", fgColor="FEF3C7"),
    "Fieldwork":   PatternFill("solid", fgColor="EDE9FE"),
    "Open":        PatternFill("solid", fgColor="FEF3C7"),
    "Overdue":     PatternFill("solid", fgColor="FEE2E2"),
    "Closed":      PatternFill("solid", fgColor="D1FAE5"),
}
SEVERITY_FILLS = {
    "High":   PatternFill("solid", fgColor="FEE2E2"),
    "Medium": PatternFill("solid", fgColor="FEF3C7"),
    "Low":    PatternFill("solid", fgColor="D1FAE5"),
}


# ── Style helpers ─────────────────────────────────────────────────────────────

def _write_header_row(ws, row: int, cols: list[str],
                      bg: str = DARK_BG, fg: str = TEXT_WHITE,
                      start_col: int = 1):
    # Reuse module-level singletons for the common default case; create new objects only for custom colors
    fill = FILL_DARK if bg == DARK_BG else PatternFill("solid", fgColor=bg)
    font = FONT_HEADER if fg == TEXT_WHITE else Font(bold=True, color=fg, size=10)
    for i, label in enumerate(cols, start=start_col):
        cell = ws.cell(row=row, column=i, value=label)
        cell.fill = fill
        cell.font = font
        cell.alignment = ALIGN_HDR_WRAP
        cell.border = THIN_BORDER


def _style_cell(cell, fill=None, font=None, align_center=False):
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    cell.alignment = ALIGN_CENTER if align_center else ALIGN_LEFT


def _set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _sheet_title(ws, title: str, subtitle: str = "", row: int = 1) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=16, color=TEXT_WHITE)
    cell.fill = FILL_DARK
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[row].height = 32

    if subtitle:
        ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=12)
        sub = ws.cell(row=row + 1, column=1, value=subtitle)
        sub.font = Font(size=10, color="9CA3AF")
        sub.fill = FILL_MID
        sub.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[row + 1].height = 18
        return row + 2
    return row + 1


def _add_bar_chart(ws, anchor: str, data_ref: Reference, cat_ref: Reference,
                   title: str, y_title: str = "Count", x_title: str = "",
                   grouping: str = None, width: int = 16, height: int = 10):
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.y_axis.title = y_title
    if x_title:
        chart.x_axis.title = x_title
    if grouping:
        chart.grouping = grouping
    chart.style = 10
    chart.width = width
    chart.height = height
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    ws.add_chart(chart, anchor)


def _get_completed_published(audits_df: pd.DataFrame):
    """Return (completed, published, core_published) sub-DataFrames."""
    completed = audits_df[audits_df["status"] == "Complete"]
    if "report_status" in completed.columns:
        published = completed[completed["report_status"] == "Published"]
    else:
        published = completed.iloc[0:0]
    if "audit_type" in published.columns:
        core = published[published["audit_type"].isin(["Owned Audit", "In-Scope AE"])]
    else:
        core = published
    return completed, published, core


# ── Sheet 1: Summary Dashboard ────────────────────────────────────────────────

def _build_summary(wb: Workbook, audits_df: pd.DataFrame,
                   issues_df: pd.DataFrame, platform: str, quarter: str):
    ws = wb.create_sheet("Summary Dashboard")
    ws.sheet_view.showGridLines = False

    next_row = _sheet_title(ws, "AuditIQ — Assurance Summary", f"{platform}  ·  {quarter}")

    completed, published, _ = _get_completed_published(audits_df)

    open_iss = issues_df[issues_df["status"].isin(["Open", "In Progress"])]
    overdue  = issues_df[issues_df["status"] == "Overdue"]

    kpis = [
        ("Total Audits",      len(audits_df), ACCENT),
        ("Completed",         len(completed), GREEN),
        ("Published Reports", len(published), "2563EB"),
        ("Open Issues",       len(open_iss),  "D97706"),
        ("Overdue Issues",    len(overdue),   "DC2626"),
    ]

    r = next_row + 1
    ws.merge_cells(start_row=r-1, start_column=1, end_row=r-1, end_column=10)
    ws.cell(row=r-1, column=1, value="KEY PERFORMANCE INDICATORS").font = Font(bold=True, size=11, color=TEXT_GRAY)

    for idx, (label, value, color) in enumerate(kpis):
        col = idx * 2 + 1
        val_cell = ws.cell(row=r, column=col, value=value)
        val_cell.font = Font(bold=True, size=22, color=color)
        val_cell.fill = FILL_KPI
        val_cell.alignment = ALIGN_CENTER
        val_cell.border = THIN_BORDER
        ws.row_dimensions[r].height = 36

        lbl_cell = ws.cell(row=r+1, column=col, value=label)
        lbl_cell.font = Font(size=9, color=TEXT_GRAY)
        lbl_cell.fill = FILL_KPI
        lbl_cell.alignment = ALIGN_CENTER
        lbl_cell.border = THIN_BORDER
        ws.row_dimensions[r+1].height = 18

    next_row = r + 3

    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=5)
    ws.cell(row=next_row, column=1, value="REPORT RATINGS").font = Font(bold=True, size=11, color=TEXT_GRAY)
    next_row += 1

    _write_header_row(ws, next_row, ["Rating", "Count", "% of Completed"])
    next_row += 1

    if "current_rating" in published.columns:
        n_pub = len(published)
        for rating in ["SAT", "RI", "UNSAT", "NA"]:
            cnt = len(published[published["current_rating"] == rating])
            pct = f"{cnt/n_pub*100:.0f}%" if n_pub else "—"
            fills = [RATING_FILLS.get(rating), None, None]
            for ci, (val, fill) in enumerate(zip([rating, cnt, pct], fills), 1):
                c = ws.cell(row=next_row, column=ci, value=val)
                _style_cell(c, fill=fill, align_center=True)
                if ci == 1:
                    c.font = RATING_FONTS.get(rating, Font())
            next_row += 1

    next_row += 1
    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=5)
    ws.cell(row=next_row, column=1, value="AUDIT TYPE BREAKDOWN").font = Font(bold=True, size=11, color=TEXT_GRAY)
    next_row += 1

    _write_header_row(ws, next_row, ["Audit Type", "Total", "Complete", "In Progress", "Fieldwork"])
    next_row += 1

    if "audit_type" in audits_df.columns:
        for atype, group in audits_df.groupby("audit_type"):
            status_counts = group["status"].value_counts()
            row_data = [
                atype, len(group),
                status_counts.get("Complete", 0),
                status_counts.get("In Progress", 0),
                status_counts.get("Fieldwork", 0),
            ]
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=next_row, column=ci, value=val)
                _style_cell(c, align_center=(ci > 1))
            next_row += 1

    _set_col_widths(ws, {"A": 22, "B": 12, "C": 16, "D": 14, "E": 14,
                         "F": 12, "G": 12, "H": 12, "I": 12, "J": 12})


# ── Sheet 2: Portfolio Overview ───────────────────────────────────────────────

def _build_portfolio(wb: Workbook, audits_df: pd.DataFrame, platform: str, quarter: str):
    ws = wb.create_sheet("Portfolio Overview")
    ws.sheet_view.showGridLines = False

    next_row = _sheet_title(ws, "Portfolio Overview", f"{platform}  ·  {quarter}")
    next_row += 1

    want_cols = [
        "audit_id", "audit_name", "audit_type", "lead_group", "region",
        "status", "rating", "current_rating", "marc_rating",
        "report_status", "issue_count", "quarter",
    ]
    display_cols = [c for c in want_cols if c in audits_df.columns]
    headers = [c.replace("_", " ").title() for c in display_cols]

    _write_header_row(ws, next_row, headers)
    ws.freeze_panes = f"A{next_row + 1}"
    next_row += 1

    color_cols = {"current_rating", "status", "rating"}
    for _, row in audits_df[display_cols].iterrows():
        for ci, col in enumerate(display_cols, 1):
            val = row[col]
            c = ws.cell(row=next_row, column=ci, value=val)
            c.border = THIN_BORDER
            c.font = FONT_BODY
            if col in color_cols:
                c.alignment = ALIGN_CENTER
                if col == "current_rating":
                    c.fill = RATING_FILLS.get(val, PatternFill())
                    c.font = RATING_FONTS.get(val, FONT_BODY)
                elif col == "status":
                    c.fill = STATUS_FILLS.get(val, PatternFill())
                elif col == "rating":
                    c.fill = SEVERITY_FILLS.get(val, PatternFill())
            else:
                c.alignment = ALIGN_LEFT
        next_row += 1

    _set_col_widths(ws, {
        "A": 12, "B": 38, "C": 16, "D": 12, "E": 12,
        "F": 14, "G": 10, "H": 14, "I": 22, "J": 16, "K": 12, "L": 12,
    })

    chart_start = next_row + 2
    status_dict = audits_df["status"].value_counts().to_dict()
    ws.cell(row=chart_start, column=1, value="Status").font = Font(bold=True)
    ws.cell(row=chart_start, column=2, value="Count").font = Font(bold=True)
    for i, (status, count) in enumerate(status_dict.items(), 1):
        ws.cell(row=chart_start + i, column=1, value=status)
        ws.cell(row=chart_start + i, column=2, value=int(count))

    n = len(status_dict)
    _add_bar_chart(
        ws, f"D{chart_start}",
        data_ref=Reference(ws, min_col=2, max_col=2, min_row=chart_start, max_row=chart_start + n),
        cat_ref=Reference(ws, min_col=1, min_row=chart_start + 1, max_row=chart_start + n),
        title="Audit Status Breakdown", x_title="Status",
    )


# ── Sheet 3: Issue Tracker ────────────────────────────────────────────────────

def _build_issues(wb: Workbook, issues_df: pd.DataFrame, platform: str, quarter: str):
    ws = wb.create_sheet("Issue Tracker")
    ws.sheet_view.showGridLines = False

    next_row = _sheet_title(ws, "Issue Tracker", f"{platform}  ·  {quarter}")
    next_row += 1

    want_cols = ["issue_id", "audit_id", "title", "severity", "status",
                 "owner", "due_date", "raised_date"]
    if "self_identified" in issues_df.columns:
        want_cols.append("self_identified")
    display_cols = [c for c in want_cols if c in issues_df.columns]
    headers = [c.replace("_", " ").title() for c in display_cols]

    _write_header_row(ws, next_row, headers)
    ws.freeze_panes = f"A{next_row + 1}"
    next_row += 1

    for _, row in issues_df[display_cols].iterrows():
        for ci, col in enumerate(display_cols, 1):
            val = row[col]
            if isinstance(val, bool):
                val = "Yes" if val else "No"
            c = ws.cell(row=next_row, column=ci, value=val)
            c.border = THIN_BORDER
            c.font = FONT_BODY
            if col == "severity":
                c.fill = SEVERITY_FILLS.get(row[col], PatternFill())
                c.alignment = ALIGN_CENTER
            elif col == "status":
                c.fill = STATUS_FILLS.get(row[col], PatternFill())
                c.alignment = ALIGN_CENTER
            else:
                c.alignment = ALIGN_LEFT
        next_row += 1

    _set_col_widths(ws, {
        "A": 12, "B": 12, "C": 42, "D": 10, "E": 14,
        "F": 16, "G": 14, "H": 14, "I": 16,
    })

    chart_start = next_row + 2
    sev_status = issues_df.groupby(["severity", "status"]).size().unstack(fill_value=0)
    n_rows = len(sev_status)
    n_cols = len(sev_status.columns)

    if n_rows and n_cols:
        ws.cell(row=chart_start, column=1, value="Severity").font = Font(bold=True)
        for ci, stat in enumerate(sev_status.columns, 2):
            ws.cell(row=chart_start, column=ci, value=stat).font = Font(bold=True)
        for ri, (sev, row) in enumerate(sev_status.iterrows(), 1):
            ws.cell(row=chart_start + ri, column=1, value=sev)
            for ci, val in enumerate(row.values, 2):
                ws.cell(row=chart_start + ri, column=ci, value=int(val))

        _add_bar_chart(
            ws, f"D{chart_start}",
            data_ref=Reference(ws, min_col=2, max_col=1 + n_cols,
                               min_row=chart_start, max_row=chart_start + n_rows),
            cat_ref=Reference(ws, min_col=1,
                              min_row=chart_start + 1, max_row=chart_start + n_rows),
            title="Issues by Severity & Status", grouping="clustered",
        )


# ── Sheet 4: Assurance Summary ────────────────────────────────────────────────

def _build_assurance_summary(wb: Workbook, audits_df: pd.DataFrame,
                              issues_df: pd.DataFrame, platform: str, quarter: str):
    ws = wb.create_sheet("Assurance Summary")
    ws.sheet_view.showGridLines = False

    next_row = _sheet_title(ws, "Assurance Summary", f"{platform}  ·  {quarter}")
    next_row += 1

    completed, published, core = _get_completed_published(audits_df)

    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=6)
    ws.cell(row=next_row, column=1, value="ASSURANCE ACTIVITIES & OUTPUT").font = Font(bold=True, color=TEXT_WHITE, size=11)
    ws.cell(row=next_row, column=1).fill = FILL_DARK
    ws.row_dimensions[next_row].height = 24
    next_row += 1

    _write_header_row(ws, next_row, ["Project Type", "Count"], bg=MID_BG)
    next_row += 1

    if "audit_type" in published.columns:
        for atype, color in [("Owned Audit", ACCENT), ("In-Scope AE", "7C3AED"), ("Indirect", TEXT_GRAY)]:
            cnt = len(published[published["audit_type"] == atype])
            c1 = ws.cell(row=next_row, column=1, value=atype)
            c2 = ws.cell(row=next_row, column=2, value=cnt)
            _style_cell(c1)
            _style_cell(c2, align_center=True)
            c2.font = Font(bold=True, color=color)
            next_row += 1

    next_row += 1
    _write_header_row(ws, next_row, ["Rating", "Count"], bg=MID_BG)
    next_row += 1

    if "current_rating" in core.columns:
        for rating in ["SAT", "RI", "UNSAT", "NA"]:
            cnt = len(core[core["current_rating"] == rating])
            c1 = ws.cell(row=next_row, column=1, value=rating)
            c2 = ws.cell(row=next_row, column=2, value=cnt)
            _style_cell(c1, fill=RATING_FILLS.get(rating))
            c1.font = RATING_FONTS.get(rating, Font())
            _style_cell(c2, align_center=True)
            next_row += 1

    next_row += 1
    _write_header_row(ws, next_row, ["MARC Rating", "Count"], bg=MID_BG)
    next_row += 1

    if "marc_rating" in core.columns:
        for mrating in ["Developed", "Substantially Developed", "Partially Developed", "Underdeveloped"]:
            cnt = len(core[core["marc_rating"] == mrating])
            c1 = ws.cell(row=next_row, column=1, value=mrating)
            c2 = ws.cell(row=next_row, column=2, value=cnt)
            _style_cell(c1)
            _style_cell(c2, align_center=True)
            next_row += 1

    next_row += 2
    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=6)
    ws.cell(row=next_row, column=1, value="ISSUES").font = Font(bold=True, color=TEXT_WHITE, size=11)
    ws.cell(row=next_row, column=1).fill = PatternFill("solid", fgColor="991B1B")
    ws.row_dimensions[next_row].height = 24
    next_row += 1

    _write_header_row(ws, next_row, ["Issue Category", "L1 (High)", "L2 (Medium)", "Total"], bg=MID_BG)
    next_row += 1

    if not issues_df.empty and "severity" in issues_df.columns:
        issues_df = issues_df.assign(
            level=issues_df["severity"].map({"High": "L1", "Medium": "L2", "Low": "L2"})
        )
        completed_ids = set(completed["audit_id"]) if "audit_id" in completed.columns else set()
        new_iss = issues_df[issues_df["audit_id"].isin(completed_ids)] if "audit_id" in issues_df.columns else issues_df

        si = new_iss[new_iss["self_identified"] == True] if "self_identified" in new_iss.columns else new_iss.iloc[0:0]
        open_iss    = issues_df[issues_df["status"].isin(["Open", "In Progress"])] if "status" in issues_df.columns else issues_df.iloc[0:0]
        overdue_iss = issues_df[issues_df["status"] == "Overdue"] if "status" in issues_df.columns else issues_df.iloc[0:0]

        for label, subset in [
            ("Newly Raised",    new_iss),
            ("Self-Identified", si),
            ("Open",            open_iss),
            ("Overdue",         overdue_iss),
        ]:
            l1 = len(subset[subset["level"] == "L1"]) if "level" in subset.columns else 0
            l2 = len(subset[subset["level"] == "L2"]) if "level" in subset.columns else 0
            for ci, val in enumerate([label, l1, l2, l1 + l2], 1):
                c = ws.cell(row=next_row, column=ci, value=val)
                _style_cell(c, align_center=(ci > 1))
            next_row += 1

    _set_col_widths(ws, {"A": 26, "B": 14, "C": 16, "D": 12, "E": 12})

    if "current_rating" in core.columns:
        r_data_row = next_row + 2
        rating_col = 4
        ws.cell(row=r_data_row, column=rating_col, value="Rating").font = Font(bold=True)
        ws.cell(row=r_data_row, column=rating_col + 1, value="Count").font = Font(bold=True)
        for i, rating in enumerate(["SAT", "RI", "UNSAT"], 1):
            ws.cell(row=r_data_row + i, column=rating_col, value=rating)
            ws.cell(row=r_data_row + i, column=rating_col + 1, value=len(core[core["current_rating"] == rating]))

        _add_bar_chart(
            ws, "D4",
            data_ref=Reference(ws, min_col=rating_col + 1, max_col=rating_col + 1,
                               min_row=r_data_row, max_row=r_data_row + 3),
            cat_ref=Reference(ws, min_col=rating_col,
                              min_row=r_data_row + 1, max_row=r_data_row + 3),
            title="Report Ratings", width=14, height=9,
        )


# ── Sheet 5: Adjustments ─────────────────────────────────────────────────────

def _build_adjustments(wb: Workbook, adjustments: list, platform: str, quarter: str):
    ws = wb.create_sheet("Adjustment Workflow")
    ws.sheet_view.showGridLines = False

    next_row = _sheet_title(ws, "Adjustment Workflow", f"{platform}  ·  {quarter}")
    next_row += 1

    if not adjustments:
        ws.cell(row=next_row, column=1, value="No adjustments for selected period.")
        return

    df = pd.DataFrame(adjustments)
    headers = [c.replace("_", " ").title() for c in df.columns]

    _write_header_row(ws, next_row, headers)
    ws.freeze_panes = f"A{next_row + 1}"
    next_row += 1

    for _, row in df.iterrows():
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=next_row, column=ci, value=str(val) if val is not None else "")
            c.border = THIN_BORDER
            c.font = FONT_BODY
            c.alignment = ALIGN_LEFT
        next_row += 1

    for i, hdr in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(hdr) + 4)


# ── Main builder ─────────────────────────────────────────────────────────────

def build_export_workbook(
    audits_df: pd.DataFrame,
    issues_df: pd.DataFrame,
    adjustments: list,
    platform: str,
    quarter: str,
    include_sheets: list[str] = None,
) -> bytes:
    if include_sheets is None:
        include_sheets = ALL_SHEETS

    wb = Workbook()
    wb.remove(wb.active)

    builders = {
        "Summary Dashboard":   lambda: _build_summary(wb, audits_df, issues_df, platform, quarter),
        "Portfolio Overview":  lambda: _build_portfolio(wb, audits_df, platform, quarter),
        "Issue Tracker":       lambda: _build_issues(wb, issues_df, platform, quarter),
        "Assurance Summary":   lambda: _build_assurance_summary(wb, audits_df, issues_df, platform, quarter),
        "Adjustment Workflow": lambda: _build_adjustments(wb, adjustments, platform, quarter),
    }
    for sheet in ALL_SHEETS:
        if sheet in include_sheets:
            builders[sheet]()

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
